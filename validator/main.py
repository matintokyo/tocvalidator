import wx
import os
from pathlib import Path
import concurrent.futures
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup
from ebooklib import epub
from . import gui
from . import replace_dict

CHAPTER_COLUMN = 2
HIERARCHY_COLUMN = 3
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


class MyFrame(gui.MyFrame):
    def __init__(self, *args, **kwds):
        super().__init__(*args, **kwds)
        self.executor = concurrent.futures.ThreadPoolExecutor()
        self.counter = 0
        self.row_count = 0
        self.epub_text = None
        self.output_folder = None
        self.BindUI()
        self.Layout()
        self.Show()

    def BindUI(self):
        self.toc_check_button.Bind(wx.EVT_BUTTON, self.on_toc_check_button_click)
        self.toc_selector.Bind(wx.EVT_FILEPICKER_CHANGED, self.onSelectToC)
        self.epub_selector.Bind(wx.EVT_FILEPICKER_CHANGED, self.onSelectEpub)

    def on_toc_check_button_click(self, event):
        input_excel = self.toc_selector.GetPath()
        wb = load_workbook(input_excel)
        ws = wb.active

        raw_epub_text = load_epub(self.epub_selector.GetPath())
        self.epub_text = makeComparable(raw_epub_text)
        output_excel = os.path.join(self.output_folder, "000_Excel_Check.xlsx")

        # Get hierarchy levels and chapter names from Excel
        hierarchy_levels = []
        chapter_names = []
        for row in ws.iter_rows(min_row=2):
            hierarchy_levels.append(row[HIERARCHY_COLUMN].value)
            chapter_names.append(row[CHAPTER_COLUMN].value)

        # Check for both types of errors
        hierarchy_errors = self.check_toc_hierarchy(hierarchy_levels)
        chapter_errors = [in_epub(name, self.epub_text) for name in chapter_names]

        has_errors = any(hierarchy_errors) or any(chapter_errors)

        if has_errors:
            # Highlight cells with errors
            for row_idx, (h_error, c_error) in enumerate(zip(hierarchy_errors, chapter_errors), start=2):
                if h_error:
                    ws.cell(row=row_idx, column=HIERARCHY_COLUMN+1).fill = YELLOW_FILL
                if c_error:
                    ws.cell(row=row_idx, column=CHAPTER_COLUMN+1).fill = YELLOW_FILL


            # Provide specific feedback about what type of errors were found
            error_message = []
            if any(hierarchy_errors):
                error_message.append("hierarchy番号")
            if any(chapter_errors):
                error_message.append("チャプター名")

            self.status_bar.SetStatusText(f"次の問題が見つかれました： {' 及び '.join(error_message)}。 エクセルを開きます。")
            try:
                wb.save(output_excel)
                os.startfile(output_excel)
            except PermissionError:
                self.status_bar.SetStatusText(
                "エラーが見つかりましたが、Excelを保存できませんでした。既に開いている場合は閉じてから再度検証してください。")
        else:
            self.status_bar.SetStatusText("Excelに問題が見つかりませんでした。")

    def check_toc_hierarchy(self, hierarchy_column):
        errors = []
        prev_level = 0

        # Check first element
        try:
            current_level = int(hierarchy_column[0]) if hierarchy_column[0] is not None else 0
        except (ValueError, TypeError):
            current_level = 0

        errors.append(current_level > 0)  # First element should be 0

        # Check remaining elements
        for level in hierarchy_column[1:]:
            try:
                current_level = int(level) if level is not None else 0
            except (ValueError, TypeError):
                current_level = 0

            errors.append((current_level - prev_level) >= 2)
            prev_level = current_level

        return errors

    def onSelectToC(self, event):
        self.output_folder = str(Path(self.toc_selector.GetPath()).parent.absolute())

        if self.epub_selector.GetPath() == "":
            epub_file = find_epub_file(self.toc_selector.GetPath())
            if epub_file is not None:
                self.epub_selector.SetPath(str(epub_file.absolute()))
                self.epub_selector.GetTextCtrl().SetInsertionPointEnd()

    def onSelectEpub(self, event):

        if self.toc_selector.GetPath() == "":
            toc_file = find_chapter_file(self.epub_selector.GetPath())
            if toc_file is not None:
                self.toc_selector.SetPath(str(toc_file.absolute()))
                self.toc_selector.GetTextCtrl().SetInsertionPointEnd()
                self.output_folder = str(Path(self.toc_selector.GetPath()).parent.absolute())


def in_epub(search, text):
    return not (makeComparable(search) in text)


def find_chapter_file(target_folder):
    parent_folder = Path(target_folder).parent.absolute()
    for file in parent_folder.glob('*.xlsx'):
        if "chapter" in os.path.basename(file):
            return file
    return None


def find_epub_file(target_folder):
    parent_folder = Path(target_folder).parent.absolute()
    for file in parent_folder.glob('*.epub'):
        return file
    return None


def makeComparable(input_text):
    if isinstance(input_text, str):
        return input_text.translate(replace_dict.REPLACE_DICT_TRANS)
    else:
        return ""

def load_epub(epub_path):
    book = epub.read_epub(epub_path)
    content = ""
    for item in book.get_items_of_type(9):
        content += item.get_body_content().decode('utf-8', errors='ignore')
    return html_to_text(content)


def html_to_text(html):
    soup = BeautifulSoup(html, features="html.parser")

    # kill all script and style elements
    for script in soup(["script", "style"]):
        script.extract()  # rip it out

    # get text
    text = soup.get_text()

    # break into lines and remove leading and trailing space on each
    lines = (line.strip() for line in text.splitlines())
    # break multi-headlines into a line each
    chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
    # drop blank lines
    text = '\n'.join(chunk for chunk in chunks if chunk)

    return text

class TocToolsApp(wx.App):
    def OnInit(self):
        self.frame = MyFrame(None, wx.ID_ANY, "")
        self.SetTopWindow(self.frame)
        self.frame.Show()
        return True

def main():
    app = TocToolsApp(0)
    app.MainLoop()

if __name__ == "__main__":
    main()
